import sys
import time
import logging
import json
import re
import sqlite3
import openpyxl
import xlsxwriter
import pandas as pd

from typing import Dict, Optional, List, Any, Union, TypeVar, Tuple, Callable
from pathlib import Path
from pandas import DataFrame
from lxml import etree

from openpyxl.styles import Font, PatternFill
#from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#Dict_Table = TypeVar('Dict_Table', bound=Dict[str, Dict[str, Any]])
#List_Str = TypeVar('List_Str', bound=Union[List[str], None])

Dict_Table = Dict[str, Dict[str, Any]]
List_Str = Union[List[str], None]

# Constants
MAX_EXCEL_COLUMN_SIZE: float = 45.0
EXCEL_COLUMN_FILTER_SIZE: float = 3.0  # to account for the filter arrow size

def elapsed_time(start_time: float) -> str:
    elapsed_time_seconds: float = time.time() - start_time
    elapsed_time_str: str = f"{elapsed_time_seconds:.1f}"  # Format elapsed time with 1 decimal places
    return elapsed_time_str

def enable_logging(m_log_file: Path, m_logging_level: int) -> None:
    # Define a custom date format without milliseconds
    custom_date_format: str = '%Y-%m-%d %H:%M:%S'

    # Configure logging with the custom date format
    logging.basicConfig(
        level=m_logging_level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        datefmt=custom_date_format,
        handlers=[
            logging.FileHandler(m_log_file),
            logging.StreamHandler()
        ]
    )

def ensure_directory_path(file_path: Path) -> None:
    # Ensure the directory of the file exists, creating it if necessary
    file_path.parent.mkdir(parents=True, exist_ok=True)

def save_to_json_dict(m_data_dict: Dict_Table, m_output_file: Path) -> None:
    logging.info(f"Saving data to json file {m_output_file}")
    ensure_directory_path(m_output_file)
    with open(m_output_file, 'w', encoding='utf-8') as f:
        json.dump(m_data_dict, f, ensure_ascii=False, indent=4)

def save_to_json_array(m_data_dict: Dict_Table, m_output_file: Path) -> None:
    logging.info(f"Saving data to json file {m_output_file}")
    ensure_directory_path(m_output_file)
    with open(m_output_file, 'w', encoding='utf-8') as f:
        json.dump(list(m_data_dict.values()), f, ensure_ascii=False, indent=4)

def merge_columns_with_same_case_insensitive_name(data_frame: DataFrame) -> DataFrame:
    merged_columns = {}
    new_columns = []

    for column in data_frame.columns:
        lowercase_column = column.lower()

        if lowercase_column not in merged_columns:
            merged_columns[lowercase_column] = [column]
            new_columns.append(column)
        else:
            original_cases = merged_columns[lowercase_column]

            # Merge the columns with the same name
            merged_column = original_cases[0]

            # replace data in `merged_column` with data in `column`
            data_frame[merged_column] = data_frame.apply(lambda row: row[merged_column] if not pd.isna(row[merged_column]) else row[column], axis=1)
            # Delete the column `column`
            del data_frame[column]

            logging.warning(f"Columns {column} have been merged to column '{merged_column}'.")

    return data_frame

def make_column_names_case_insensitive_unique_frame(data_frame: DataFrame) -> List[str]:
    seen_columns = {}
    new_columns = []

    for column in data_frame.columns:
        lowercase_column = column.lower()

        if lowercase_column not in seen_columns:
            seen_columns[lowercase_column] = [column]
            new_columns.append(column)
        else:
            original_cases = seen_columns[lowercase_column]
            unique_name = None
            i = 1
            while unique_name is None or unique_name in data_frame.columns:
                unique_name = f"{column}_{i}"
                i += 1

            original_cases.append(unique_name)
            new_columns.append(unique_name)

            logging.warning(f"Column '{column}' has been renamed to '{unique_name}'")

    return new_columns

def save_data_dict_to_sqlite(data_dict: Dict_Table, db_file: Path, db_table: str, columns_order: Optional[List_Str] = None) -> None:
    # Create a DataFrame from the data dictionary
    data_frame = DataFrame.from_dict(data_dict, orient="index")
    save_data_frame_to_sqlite(data_frame, db_file, db_table, columns_order)
        
def save_data_frame_to_sqlite(data_frame: DataFrame, db_file: Path, db_table: str, columns_order: Optional[List_Str] = None) -> None:
    logging.info(f"Saving data to SQLite tables in database '{db_file}'; table '{db_table}'")

    # Ensure the directory path exists
    ensure_directory_path(db_file)

    original_columns = list(data_frame.columns)
    data_frame = merge_columns_with_same_case_insensitive_name(data_frame)

    # Create a SQLite connection and save the DataFrame to the SQLite table
    with sqlite3.connect(db_file) as conn:
        data_frame.to_sql(db_table, conn, if_exists='replace', index=False)
        conn.commit() # not needed but safe ...
    # No need to explicitly commit or close the connection

def column_exists(cursor: sqlite3.Cursor, db_table: str, col_name: str) -> bool:
    cursor.execute(f'PRAGMA table_info("{db_table}")')
    existing_columns = [column[1].lower() for column in cursor.fetchall()]
    return col_name.lower() in existing_columns

def save_data_dict_to_sqlite_direct(data_dict: Dict_Table, db_file: Path, db_table: str, columns_order: Optional[List_Str] = None) -> None:
    logging.info(f"Saving data to SQLite table '{db_table}' in database '{db_file}' (sqlite3 lib)")

    # Create a connection to the SQLite database file
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()

    # Check if the table exists, and if not, create it
    cursor.execute(f'CREATE TABLE IF NOT EXISTS "{db_table}" (id INTEGER PRIMARY KEY AUTOINCREMENT)')

    # Get a comprehensive, prioritized, and ordered list of all column names
    column_names: List[str] = []

    # Add columns in the specified order if provided
    if columns_order is not None:
        column_names.extend(column for column in columns_order if column not in column_names)

    # Add remaining columns from data_dict with non-empty values
    columns_with_non_empty_values = set()
    for inner_dict in data_dict.values():
        for key, value in inner_dict.items():
            if value != '':
                columns_with_non_empty_values.add(key)

    # Convert columns_order to a set for efficient membership testing
    columns_order_set = set(columns_order) if columns_order else set()

    # Add the columns with non-empty values to the column_names list, preserving order
    column_names.extend(col for col in columns_with_non_empty_values if col not in columns_order_set)

    # Create columns in the table based on the final order if they don't exist
    for col_name in column_names:
        if not column_exists(cursor, db_table, col_name):
            cursor.execute(f'ALTER TABLE "{db_table}" ADD COLUMN "{col_name}" TEXT')

    # Insert data into the table (avoid inserting empty columns)
    for row_data in data_dict.values():
        # Calculate column names and non-empty values dynamically
        column_names = list(row_data.keys())
        non_empty_columns = [col for col, value in row_data.items() if value != '']
        non_empty_values = [row_data[col] for col in non_empty_columns]

        if any(non_empty_values):
            quoted_columns = [f'"{col}"' for col in non_empty_columns]
            # Prepare the SQL statement with placeholders for each non-empty column
            sql = f'INSERT INTO "{db_table}" ({", ".join(quoted_columns)}) ' \
                    f'VALUES ({", ".join(["?" for _ in non_empty_values])})'

            # Execute the SQL statement with the corresponding non-empty values
            cursor.execute(sql, non_empty_values)

    # Commit the changes and close the connection
    conn.commit()
    conn.close()

def save_to_excel_with_xlsxwriter_direct(data_dict: Dict_Table, output_file: Path, output_sheet: str, columns_order: Optional[List_Str] = None) -> None:
    logging.info(f"Saving {output_sheet} sheet to excel workbook {output_file}, (xlsxwriter library)")
    start_time = time.time()

    output_file_path = Path(output_file)
    
    # Ensure the directory path exists
    ensure_directory_path(output_file_path)

    # Create a new XlsxWriter workbook and add a worksheet.
    workbook = xlsxwriter.Workbook(output_file_path)
    worksheet = workbook.add_worksheet(output_sheet)

    # Disable the interpretation of cells starting with '=' as formulas
    workbook.strings_to_formulas = False

    # Get a comprehensive prioritized and ordered list of all column names in the two level data_dict:
    column_names : List[str] = []
    if columns_order is not None:
        column_names.extend(column for column in columns_order if column not in column_names)
    for inner_dict in data_dict.values():
        column_names.extend(key for key in inner_dict.keys() if key not in column_names)

    # Use a list to store maximum column widths, initialized with zeros
    max_column_widths: List[float] = [0.0] * len(column_names)

    # Define a cell format for the header row.
    header_format = workbook.add_format({'bold': True, 'bg_color': 'cyan'})

    # Write the header row (column names).
    row_num = 0
    for col_num, column_name in enumerate(column_names, start=0):
        worksheet.write_string(row_num, col_num, column_name, header_format)
        max_column_widths[col_num] = len(column_name) + EXCEL_COLUMN_FILTER_SIZE

    # Create a custom text format to set the cell type to text (when needed)
    #text_format = workbook.add_format({'num_format': '@'})

    # Write the data from data_dict to the worksheet
    row_num = 1
    for outer_key, inner_dict in data_dict.items():
        for col_num, column_name in enumerate(column_names, start=0):
            if column_name in inner_dict:                       # 1. check if column exist in the record
                cell_value = inner_dict.get(column_name, '')    # Retrieve the value
                if cell_value != '':                            # 2. check if value is not empty in that column!

                    # Check if cell_value starts with '='
                    #if cell_value.startswith('='):
                    #    # Set the cell type to text to prevent formula interpretation
                    #    worksheet.write_string(row_num, col_num, str(cell_value), text_format)
                    #else:
                    #    worksheet.write_string(row_num, col_num, str(cell_value))

                    worksheet.write_string(row_num, col_num, str(cell_value))
                    # fill-in max_column_width in parallel
                    max_column_widths[col_num] = max(max_column_widths[col_num], len(str(cell_value)))
        row_num += 1

    # Re-size columns but not exceeding MAX_EXCEL_COLUMN_SIZE
    col_num = 0
    for max_width in max_column_widths:
        max_width = min(max_width, MAX_EXCEL_COLUMN_SIZE)
        worksheet.set_column(col_num, col_num, max_width)
        col_num += 1
    # Freeze the first row
    worksheet.freeze_panes(1, 0)

    # Enable data filtering on the first row
    worksheet.autofilter(0, 0, row_num-1, col_num-1)

    # Save the workbook
    workbook.close()

    #logging.info(f"Saving completed in {elapsed_time(start_time)} (xlsxwriter library)")

def save_to_excel_with_openpyxl_direct(data_dict: Dict_Table, output_file: Path, output_sheet: str, columns_order: Optional[List_Str] = None) -> None:
    
    logging.info(f"Save to excel workbook {output_file}, sheet: {output_sheet} (openpyxl library)")
    start_time = time.time()
    
    # Ensure the directory path exists
    ensure_directory_path(output_file)
    
    if output_file.is_file():
        workbook = openpyxl.load_workbook(output_file)
    else:
        workbook = Workbook()
        # Access the default first sheet (index 0) and rename it to 'output_sheet'
        sheet = workbook.active
        assert sheet is not None
        sheet.title = output_sheet
    
    if output_sheet in workbook.sheetnames:
        sheet  = workbook[output_sheet]
    else:
        sheet = workbook.create_sheet(title=output_sheet)
        
    # Create  ap prioritized list for column names
    column_names : List[str] = []
    # First, add the priority columns to the list
    if columns_order:
        for column in columns_order:
            if column not in column_names:
                column_names.append(column)
    # Then, add the remaining columns to the list
    for inner_dict in data_dict.values():
        for key in inner_dict.keys():
            if key not in column_names:
                column_names.append(key)

    # Use a list to store maximum column widths, initialized with zeros
    max_column_widths: List[float] = [0] * len(column_names)

    # formats for the header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="00CCFFFF", end_color="00CCFFFF", fill_type="solid")

    # Write the headers, set data type and formatting
    row_num = 1
    for col_num, column_name in enumerate(column_names, start=1):
        cell = sheet.cell(row=row_num, column=col_num, value=column_name)
        #cell.data_type = "s"
        cell.font = header_font
        cell.fill = header_fill
        max_column_widths[col_num - 1] = len(column_name) + EXCEL_COLUMN_FILTER_SIZE


    # Write the data from data_dict to the worksheet
    row_num = 2
    for outer_key, inner_dict in data_dict.items():
        for col_num, column_name in enumerate(column_names, start=1):
            if column_name in inner_dict: # 1. check if column exist in the record
                cell_value = inner_dict.get(column_name, '')  # Retrieve the value
                if cell_value != '': # 2. check if value is not empty in that column!
                    cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
                    if str(cell_value).startswith('='):
                        cell.data_type = 's'  # Treat it as text
                    max_column_widths[col_num - 1] = max(max_column_widths[col_num - 1], len(str(cell_value)))
        row_num += 1

    # Re-size columns but not exceeding MAX_EXCEL_COLUMN_SIZE
    for col_num, max_width in enumerate(max_column_widths, start=1):
        column_letter = get_column_letter(col_num)
        column_dimension = sheet.column_dimensions[column_letter]
        column_dimension.width = min(max_width, MAX_EXCEL_COLUMN_SIZE)

    # Freeze the panes just below the headers
    sheet.freeze_panes = "A2"
        
    # Enable data filtering on the first row
    sheet.auto_filter.ref = sheet.dimensions

    # Save the workbook
    workbook.save(output_file)
    
    # Done
    #logging.info(f"Save to excel completed in {elapsed_time(start_time)} (openpyxl library)")
    
def sort_data_dict_columns(m_data_dict: Dict_Table, m_columns_order: List_Str) -> Dict_Table:
    if not m_columns_order:
        return m_data_dict
    sorted_data_dict: Dict_Table = {}
    for key, values in m_data_dict.items():
        sorted_values = {col: values[col] for col in m_columns_order if col in values}
        # Include all other columns not in column_order
        sorted_values.update(values)
        sorted_data_dict[key] = sorted_values
    return sorted_data_dict


# Define the Translator class to handle language files and translations
class Translator:
    def __init__(self):
        self.translations = {}  # Initialize translations
        self.languages = {} # Initialize languages

    def load_translations(self, _translation_languages, _unpacked_data_folder):
        self.languages = _translation_languages
        try:
            for language in _translation_languages:
                language_file = language + ".loca.xml"
                language_path = _unpacked_data_folder / language / "Localization" / language / language_file
                logging.info(f"TRANSLATIONS - {language} language loading")

                # parse the XML file using lxml
                parser = etree.XMLParser()
                tree = etree.parse(language_path, parser=parser)
                root = tree.getroot()
                
                # Initialize data for DataFrame
                data = {
                    'UID': [],
                    'Version': [],
                    'Text': []
                }

                # Populate data from XML
                for content in root.findall('.//content'):
                    contentuid = content.get('contentuid')
                    version = content.get('version')
                    localized_text = content.text
                    data['UID'].append(contentuid)
                    data['Version'].append(version)
                    data['Text'].append(localized_text)

                # Create DataFrame and store it in the dictionary
                self.translations[language] = {}
                for i in range(len(data['UID'])):
                    uid = data['UID'][i]
                    version = data['Version'][i]
                    text = data['Text'][i]
                    if uid not in self.translations[language]:
                        self.translations[language][uid] = {}
                    self.translations[language][uid][version] = text

                num_rows = len(data['UID'])
                logging.info(
                    f"TRANSLATIONS - {language} language loaded: {num_rows} rows")

        except Exception as e:
            logging.critical(
                f"An error occurred while processing {_translation_languages} language file: {e}")
            sys.exit(1)

    def get_text(self, language, UID, Version):
        if language in self.translations:
            language_data = self.translations[language]
            if UID in language_data and Version in language_data[UID]:
                return language_data[UID][Version]
        return None

# Create an instance of the Translator class (global variable - singleton)
global translator
translator = Translator()

# Regular expression pattern for matching translation handles (mind the g in 0-9a-g!)
add_localized_text_pattern = re.compile(r'^h[0-9a-g]{36}$')

def add_localized_text(m_current_record: Dict[str, Any], id: str, m_localized_handle: str, m_localized_version: str) -> None:
    if m_localized_handle == "ls::TranslatedStringRepository::s_HandleUnknown;0":
        return

    for language in translator.languages:
        LocalizedKey = f"{id}{language}"

        if not add_localized_text_pattern.match(m_localized_handle):
            error_text = f"Localized Text for id \"{id}\", handle \"{m_localized_handle}\", version \"{m_localized_version}\" in \"{language}\" not found"
            logging.debug(error_text)
            continue

        LocalizedValue = translator.get_text(language, m_localized_handle, m_localized_version)
        if LocalizedValue:
            m_current_record[LocalizedKey] = LocalizedValue
        else:
            error_text = f"Localized Text for id \"{id}\", handle \"{m_localized_handle}\", version \"{m_localized_version}\" in \"{language}\" not found"
            if id in ("DisplayName", "Description"):
                # NOTE: usually not an issue! probably an updated version will be found later
                logging.debug(error_text)
            else:
                # NOTE: NA for TechnicalDescription, GameMasterSpawnSubSection, OnUseDescription, ShortDescription, etc
                logging.debug(error_text)

def process_lsx_files(m_unpacked_data_folder: Path,
                      m_folder: Path,
                      m_file_pattern: str,
                      m_id_name: str,
                      m_key_name: str,
                      m_data_dict: Dict_Table,
                      m_children: Optional[str] = None,
                      m_post_process_func: Optional[Callable[..., None]] = None) -> Tuple[int, int]:

    lsx_files = list(m_folder.rglob(m_file_pattern))

    if not lsx_files:
        return 0, 0  # No matching files found, nothing to process

    relative_path = m_folder.relative_to(m_unpacked_data_folder)

    if m_file_pattern != '_merged.lsf.lsx':  # we are not processing merged files
        len_lsx_files = len(lsx_files)
        logging.info(f"{relative_path} - Processing {len_lsx_files} files")

    overwritten_nodes = 0
    lsx_files_count = 0
    for lsx_files_count, lsx_file in enumerate(lsx_files, start=1):
        parser = etree.XMLParser()
        try:
            xml_document = etree.parse(str(lsx_file),parser=parser)
        except etree.XMLSyntaxError as e:
            logging.error(f"Error parsing {lsx_file}: {e}")
            continue

        ID_nodes = xml_document.findall(f".//node[@id='{m_id_name}']")
        node_count = len(ID_nodes)  # Get the count of nodes
        #1 relative_file = lsx_file.relative_to(m_unpacked_data_folder)
        #1 if m_file_pattern == '_merged.lsf.lsx':
        #1    logging.info(f"{relative_file} - Processing {node_count} '{m_id_name}' nodes")

        for ID_node in ID_nodes:
            record: Dict[str, Any] = {}  # Initialize the record dictionary here

            # attribute_nodes = ID_node.findall("./attribute")
            # for attr_node in attribute_nodes:
            for attr_node in ID_node.iterfind("./attribute"):
                attribute_id = attr_node.get("id")
                attribute_type = attr_node.get("type")
                if attribute_type == "TranslatedString":
                    attribute_handle = attr_node.get("handle")
                    attribute_version = attr_node.get("version")
                    #if attribute_handle != "ls::TranslatedStringRepository::s_HandleUnknown":
                    if not attribute_handle.startswith("ls:"):
                        record[attribute_id] = f"{attribute_handle};{attribute_version}"
                        add_localized_text(record, attribute_id, attribute_handle, attribute_version)
                else:
                    attribute_value = attr_node.get("value")
                    if attribute_value:
                        record[attribute_id] = attribute_value

            if m_children:
                child_nodes = ID_node.findall(f".//node[@id='{m_children}']")
                childs = ", ".join(child_node.find(".//attribute[@id='Name']").attrib["value"] for child_node in child_nodes)
                if childs:
                    record[m_children] = childs

            record['RootFolder'] = str(lsx_file.parent.relative_to(m_unpacked_data_folder))
            record['RootFile'] = str(lsx_file.name)

            if callable(m_post_process_func):
                m_post_process_func(record, m_key_name)

            if m_key_name in record:
                new_key = record[m_key_name]
                if new_key in m_data_dict:
                    overwritten_nodes += 1
                    logging.debug(f"Overwriting duplicate entry {new_key}")
                m_data_dict[new_key] = record
            else:
                lsx_file_name = str(lsx_file.relative_to(m_unpacked_data_folder))
                logging.error(f"Missing {m_key_name} for {lsx_file_name}")

        #1 if overwritten_nodes:
        #1    logging.debug(f"{relative_file} - {overwritten_nodes} overwritten nodes!")

    logging.debug(f"{relative_path} - Processed: {lsx_files_count} files")
    return lsx_files_count, overwritten_nodes


def data_frame_from_dict(data_dict:Dict_Table, columns_order:List_Str) -> DataFrame:
    # Create a DataFrame from the data dictionary
    data_frame = pd.DataFrame.from_dict(data_dict, orient="index")

    # Initialize an empty list for the reordered columns
    reordered_columns = []

    # Add columns from columns_order if they exist in the DataFrame
    for col in columns_order:
        if col in data_frame.columns:
            reordered_columns.append(col)
    # Add all remaining columns from the DataFrame that were not explicitly in columns_order
    remaining_columns = [
        col for col in data_frame.columns if col not in reordered_columns]
    reordered_columns.extend(remaining_columns)

    # Create the reordered DataFrame
    data_frame = data_frame[reordered_columns]
    
    return data_frame

class ExcelWriter:
    def __init__(self):
        self.workbook: Optional[xlsxwriter.Workbook] = None
        self.file_path: Path = None
        self.total_time: float = 0.0
        
    def open_workbook(self, output_file: Path) -> None:
        self.file_path = Path(output_file)

        # Ensure the directory path exists
        self.file_path.parent.mkdir(parents=True, exist_ok=True)

        # Create a new XlsxWriter workbook and add a worksheet.
        self.workbook = xlsxwriter.Workbook(self.file_path)

        # Disable the interpretation of cells starting with '=' as formulas
        self.workbook.strings_to_formulas = False
        logging.info(f"Workbook {self.file_path}: Created")

    def add_sheet(self, data_dict: Dict_Table, output_sheet: str, columns_order: Optional[List_Str] = None) -> None:
        start_time: float = time.time()
        if self.workbook is None:
            raise ValueError("Workbook is not open. Call open_workbook() before adding sheets.")

        worksheet = self.workbook.add_worksheet(output_sheet)

        # Get a prioritized and ordered list of all column names in the two-level data_dict:
        column_names: List[str] = []
        if columns_order is not None:
            column_names.extend(column for column in columns_order if column not in column_names)
        for inner_dict in data_dict.values():
            column_names.extend(key for key in inner_dict.keys() if key not in column_names)

        # Use a list to store maximum column widths, initialized with zeros
        max_column_widths: List[float] = [0.0] * len(column_names)

        # Define a cell format for the header row.
        header_format = self.workbook.add_format({'bold': True, 'bg_color': 'cyan'})

        # Write the header row (column names).
        row_num = 0
        for col_num, column_name in enumerate(column_names, start=0):
            worksheet.write_string(row_num, col_num, column_name, header_format)
            max_column_widths[col_num] = len(column_name) + EXCEL_COLUMN_FILTER_SIZE

        # Create a custom text format to set the cell type to text (when needed)
        #text_format = workbook.add_format({'num_format': '@'})

        # Write the data from data_dict to the worksheet
        row_num = 1
        for outer_key, inner_dict in data_dict.items():
            for col_num, column_name in enumerate(column_names, start=0):
                if column_name in inner_dict:                       # Check if column exist in the record
                    cell_value = inner_dict.get(column_name, '')    # Retrieve the value
                    if cell_value != '':                            # check if value is not empty in that column!
                        # write the cell value
                        worksheet.write_string(row_num, col_num, str(cell_value))
                        # fill-in max_column_width in parallel
                        max_column_widths[col_num] = max(max_column_widths[col_num], len(str(cell_value)))
            row_num += 1

        col_num = 0
        for max_width in max_column_widths:
            max_width = min(max_width, MAX_EXCEL_COLUMN_SIZE)
            worksheet.set_column(col_num, col_num, max_width)
            col_num += 1

        # Freeze the first row
        worksheet.freeze_panes(1, 0)
        
        # Enable data filtering on the first row
        worksheet.autofilter(0, 0, row_num - 1, col_num - 1)
        logging.info(f"Workbook {self.file_path}: Sheet {output_sheet} Added ({elapsed_time(start_time)})")
        self.total_time += (time.time() - start_time)
    def close_workbook(self) -> None:
        if self.workbook is None:
            raise ValueError("Workbook is not open. Call open_workbook() before closing.")

        # Save the workbook
        self.workbook.close()
        self.workbook = None
        logging.info(f"Workbook {self.file_path}: Saved ({self.total_time:.1f})")

# The following code will only run if this module is executed directly
if __name__ == "__main__":
    print("This is a library module and should not be executed directly.")
    print("It is ment to be used with `bg3data_dumps.py` main module.")
